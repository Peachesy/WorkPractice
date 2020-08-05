import openpyxl  # 对xlsx格式的excel进行读取和编辑

# 从excel读取数据后再向excel写入数据
wb = openpyxl.load_workbook('test.xlsx')  # 打开test.xlsx文件，获取workbook对象
sheets = wb.sheetnames  # 从workbook中获取一个表单对象
print(sheets,type(sheets))

recording = wb.create_sheet('Recording')  # 创建一个表单叫recording 但在我的excel文件中并没有新增一个叫recording的sheet？？？？？？？？？？？？？？
print(wb.sheetnames)

# 获取指定的表单worksheet
buglist = wb['Bug List']  # get_sheet_by_name()是不推荐的用法
print(buglist,type(buglist))
# for sheet in buglist:
#     print((sheet))
# ?????我也不知道这段for循环什么意思？？？？？

print(buglist['A1'])
print(buglist['A1'].value)

ws=wb.active
print(ws)
print('第一行第二列：',ws.cell(row=1,column=2).value)
# 打印Test point列的所有数据
for i in ws['B']:
    print(i.value)
    # print(i,ws.cell(row=i,column=2).value)

ts = wb['Test Cases']
# print("Test Link is :")
for i in ts['C']:
    print(i.value)

# 向worksheet中的result列写入数据
# for i in ts['D']:  # i是一个cell对象,表示一个单元格,所以ts.cell(int(i),4).value = 'Pass'这种用法是错的
# for i in range(1,7):
#     ts.cell(i,4).value = 'Pass'

ts.cell(3,4).value = 7


